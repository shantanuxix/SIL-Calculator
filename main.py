import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkinter.simpledialog import askinteger
import openpyxl

component_data = []
serial_number = ""
components_added = False
window = tk.Tk()
window.title('Reliability Calculator')
window.geometry("450x600")

# Create a scrollable canvas
canvas = tk.Canvas(window)
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Add a vertical scrollbar to the canvas
scrollbar = ttk.Scrollbar(window, orient=tk.VERTICAL, command=canvas.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Configure the canvas to use the scrollbar
canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))

# Create a frame inside the canvas for the main window
main_frame = ttk.Frame(canvas)
canvas.create_window((0, 0), window=main_frame, anchor='nw')


# Phase 1 - Component Input Frame
component_input_frame = ttk.LabelFrame(main_frame, text='Component Input')
component_input_frame.grid(row=0, column=0, padx=10, pady=10, sticky='w')

# Create input labels and entry fields
label_tag_no = ttk.Label(component_input_frame, text='TAG NOS:')
label_tag_no.grid(row=0, column=0, padx=5, pady=5, sticky='w')
entry_tag_no = ttk.Entry(component_input_frame)
entry_tag_no.grid(row=0, column=1, padx=5, pady=5)

label_component = ttk.Label(component_input_frame, text='COMPONENT/INSTRUMENT:')
label_component.grid(row=1, column=0, padx=5, pady=5, sticky='w')
entry_component = ttk.Entry(component_input_frame)
entry_component.grid(row=1, column=1, padx=5, pady=5)

label_make = ttk.Label(component_input_frame, text='MAKE:')
label_make.grid(row=2, column=0, padx=5, pady=5, sticky='w')
entry_make = ttk.Entry(component_input_frame)
entry_make.grid(row=2, column=1, padx=5, pady=5)

label_model = ttk.Label(component_input_frame, text='MODEL:')
label_model.grid(row=3, column=0, padx=5, pady=5, sticky='w')
entry_model = ttk.Entry(component_input_frame)
entry_model.grid(row=3, column=1, padx=5, pady=5)

label_reliability_data_ref = ttk.Label(component_input_frame, text='Reliability Data Reference:')
label_reliability_data_ref.grid(row=4, column=0, padx=5, pady=5, sticky='w')
entry_reliability_data_ref = ttk.Entry(component_input_frame)
entry_reliability_data_ref.grid(row=4, column=1, padx=5, pady=5)

label_lambda_sd = ttk.Label(component_input_frame, text='λsd:')
label_lambda_sd.grid(row=5, column=0, padx=5, pady=5, sticky='w')
entry_lambda_sd = ttk.Entry(component_input_frame)
entry_lambda_sd.grid(row=5, column=1, padx=5, pady=5)

label_lambda_su = ttk.Label(component_input_frame, text='λsu:')
label_lambda_su.grid(row=6, column=0, padx=5, pady=5, sticky='w')
entry_lambda_su = ttk.Entry(component_input_frame)
entry_lambda_su.grid(row=6, column=1, padx=5, pady=5)

label_lambda_dd = ttk.Label(component_input_frame, text='λdd:')
label_lambda_dd.grid(row=7, column=0, padx=5, pady=5, sticky='w')
entry_lambda_dd = ttk.Entry(component_input_frame)
entry_lambda_dd.grid(row=7, column=1, padx=5, pady=5)

label_lambda_du = ttk.Label(component_input_frame, text='λdu:')
label_lambda_du.grid(row=8, column=0, padx=5, pady=5, sticky='w')
entry_lambda_du = ttk.Entry(component_input_frame)
entry_lambda_du.grid(row=8, column=1, padx=5, pady=5)

label_configuration = ttk.Label(component_input_frame, text='CONFIGURATION:')
label_configuration.grid(row=9, column=0, padx=5, pady=5, sticky='w')
entry_configuration = ttk.Entry(component_input_frame)
entry_configuration.grid(row=9, column=1, padx=5, pady=5)

label_proof_test_interval = ttk.Label(component_input_frame, text='PROOF TEST INTERVAL T1 (hrs):')
label_proof_test_interval.grid(row=10, column=0, padx=5, pady=5, sticky='w')
entry_proof_test_interval = ttk.Entry(component_input_frame)
entry_proof_test_interval.grid(row=10, column=1, padx=5, pady=5)

label_mttr = ttk.Label(component_input_frame, text='MTTR (hrs):')
label_mttr.grid(row=11, column=0, padx=5, pady=5, sticky='w')
entry_mttr = ttk.Entry(component_input_frame)
entry_mttr.grid(row=11, column=1, padx=5, pady=5)



label_remark = ttk.Label(component_input_frame, text='Remarks:')
label_remark.grid(row=12, column=0, padx=5, pady=5, sticky='w')
entry_remark = ttk.Entry(component_input_frame)
entry_remark.grid(row=12, column=1, padx=5, pady=5)




# Calculation Option Frame
calculation_option_frame = ttk.LabelFrame(main_frame, text='Calculation Option')
calculation_option_frame.grid(row=1, column=0, padx=10, pady=10, sticky='w')

def choose_calculation_option():
    calculation_option = calculation_option_combobox.get()
    if calculation_option == 'Calculate Values':
        calculate_values()

calculation_option_label = ttk.Label(calculation_option_frame, text='Calculation Option:')
calculation_option_label.grid(row=0, column=0, padx=5, pady=5)

calculation_option_combobox = ttk.Combobox(calculation_option_frame, values=['Calculate Values'])
calculation_option_combobox.grid(row=0, column=1, padx=5, pady=5)
calculation_option_combobox.current(0)

calculation_option_button = ttk.Button(calculation_option_frame, text='Choose', command=choose_calculation_option)
calculation_option_button.grid(row=0, column=2, padx=5, pady=5)



# Add Component Button
add_component_button = ttk.Button(main_frame, text='ADD COMPONENT')
add_component_button.grid(row=3, column=0, pady=10)

# Generate Report Button
generate_report_button = ttk.Button(main_frame, text='GENERATE REPORT')
generate_report_button.grid(row=4, column=0, pady=10)

last_loaded_row = 2 

# Load Annexure Button
def load_annexure():
    global last_loaded_row
    file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx'), ('All Files', '*.*')])
    if file_path:
        selected_row = askinteger("Select Row", "Enter the row number to load data from:", minvalue=2)
        if selected_row is not None:
            annexure_data = load_annexure_data(file_path, selected_row)
            if annexure_data:
                last_loaded_row = selected_row + 1
                populate_input_fields(annexure_data)
                messagebox.showinfo('Annexure Loaded', 'Annexure data loaded successfully.')
            else:
                messagebox.showwarning('Invalid Row', 'Invalid row number. Please select a valid row from the annexure.')

load_annexure_button = ttk.Button(main_frame, text='LOAD ANNEXURE', command=load_annexure)
load_annexure_button.grid(row=5, column=0, pady=10)

# Load Annexure Data Function
def load_annexure_data(file_path, row_number):
    annexure_data = {}
    wb = openpyxl.load_workbook(file_path)
    sheet_name = 'Annexure-1'
    if not sheet_name:
        return None

    # Get the selected sheet
    sheet = wb[sheet_name]

    # Check if the sheet has data starting from row 2
    if sheet.max_row >= row_number:
        annexure_data['TAG NOS'] = sheet.cell(row=row_number, column=2).value
        annexure_data['COMPONENT/INSTRUMENT'] = sheet.cell(row=row_number, column=3).value
        annexure_data['MAKE'] = sheet.cell(row=row_number, column=4).value
        annexure_data['MODEL'] = sheet.cell(row=row_number, column=5).value
        annexure_data['Reliability Data Reference'] = sheet.cell(row=row_number, column=6).value
        annexure_data['λsd'] = sheet.cell(row=row_number, column=7).value
        annexure_data['λsu'] = sheet.cell(row=row_number, column=8).value
        annexure_data['λdd'] = sheet.cell(row=row_number, column=9).value
        annexure_data['λdu'] = sheet.cell(row=row_number, column=10).value
        annexure_data['CONFIGURATION'] = sheet.cell(row=row_number, column=11).value
        annexure_data['PROOF TEST INTERVAL T1 (hrs)'] = sheet.cell(row=row_number, column=12).value
        annexure_data['MTTR (hrs)'] = sheet.cell(row=row_number, column=13).value
    else:
        messagebox.showerror('Error', 'Invalid row number. The selected row does not exist in the sheet.')
        return None
    return annexure_data

# Populate Input Fields Function
def populate_input_fields(annexure_data):
    entry_tag_no.delete(0, tk.END)
    entry_tag_no.insert(0, annexure_data.get('TAG NOS', ''))
    entry_component.delete(0, tk.END)
    entry_component.insert(0, annexure_data.get('COMPONENT/INSTRUMENT', ''))
    entry_make.delete(0, tk.END)
    entry_make.insert(0, annexure_data.get('MAKE', ''))
    entry_model.delete(0, tk.END)
    entry_model.insert(0, annexure_data.get('MODEL', ''))
    entry_reliability_data_ref.delete(0, tk.END)
    entry_reliability_data_ref.insert(0, annexure_data.get('Reliability Data Reference', ''))

    
    # Handle lambda values
    entry_lambda_sd.delete(0, tk.END)
    lambda_sd = annexure_data.get('λsd')
    if lambda_sd is not None:
        entry_lambda_sd.insert(tk.END, str(lambda_sd))

    entry_lambda_su.delete(0, tk.END)
    lambda_su = annexure_data.get('λsu')
    if lambda_su is not None:
        entry_lambda_su.insert(tk.END, str(lambda_su))

    entry_lambda_dd.delete(0, tk.END)
    lambda_dd = annexure_data.get('λdd')
    if lambda_dd is not None:
        entry_lambda_dd.insert(tk.END, str(lambda_dd))

    entry_lambda_du.delete(0, tk.END)
    lambda_du = annexure_data.get('λdu')
    if lambda_du is not None:
        entry_lambda_du.insert(tk.END, str(lambda_du))


    entry_configuration.delete(0, tk.END)
    entry_configuration.insert(tk.END, annexure_data.get('CONFIGURATION', ''))
    entry_proof_test_interval.delete(0, tk.END)
    entry_proof_test_interval.insert(0, annexure_data.get('PROOF TEST INTERVAL T1 (hrs)', ''))
    entry_mttr.delete(0, tk.END)
    entry_mttr.insert(0, annexure_data.get('MTTR (hrs)', ''))

    

    load_annexure_button = ttk.Button(main_frame, text='LOAD ANNEXURE', command=load_annexure)
load_annexure_button.grid(row=5, column=0, pady=10)

def calculate_values():
    tag_no = entry_tag_no.get()
    component = entry_component.get()
    make = entry_make.get()
    model = entry_model.get()
    reliability_data_ref = entry_reliability_data_ref.get()
    λsd = entry_lambda_sd.get()
    λsu = entry_lambda_su.get()
    λdd = entry_lambda_dd.get()
    λdu = entry_lambda_du.get()
    configuration = entry_configuration.get()
    proof_test_interval = entry_proof_test_interval.get()
    mttr = entry_mttr.get()
    
    remark= entry_remark.get()
    

    # Perform the calculations
    sff = (1 - float(λdu) /(float(λsd) + float(λsu) + float(λdd) + float(λdu))) * 100
    dc = float(λdd) / (float(λdd) + float(λdu))
    dff = 1 - sff / 100
    β = 0.1
    βd = 0.05
    λ = float(λsd) + float(λsu) + float(λdd) + float(λdu)
    mtbf = 1 / λ
    λd = λ * dff
    λdd = λd * dc
    λdu = λd * (1 - dc)
    tde = (float(λdu) * ((float(proof_test_interval) / 2) + float(mttr))) + (float(λdd) * float(mttr)) / λd
    tse = (float(λdu) * ((float(proof_test_interval) / 3) + float(mttr))) + (float(λdd) * float(mttr)) / λd
    tce = None
    if entry_configuration.get()== '2oo3':
        pfd = (6 * ((1 - float(βd)) * float(λdd) + (1 - float(β)) * float(λdu)) ** 2 * tde * tse) + \
          (float(βd) * float(λdd) * float(mttr)) + \
          (float(β) * float(λdu) * ((float(proof_test_interval) / 2) + float(mttr)))
    elif entry_configuration.get()== '2oo2':
        pfd= 2*λd*tce
    elif entry_configuration.get()== '1oo1':
        pfd=λd*tde
    elif entry_configuration.get()== '1oo2':
        pfd= (2 * ((1 - float(βd)) * float(λdd) + (1 - float(β)) * float(λdu)) ** 2 * tde * tse) + \
          (float(βd) * float(λdd) * float(mttr)) + \
          (float(β) * float(λdu) * ((float(proof_test_interval) / 2) + float(mttr)))

    # Display the calculated values in the output text
    output_text.delete('1.0', tk.END)
    output_text.insert(tk.END, f'SFF (in %): {sff}\n')
    output_text.insert(tk.END, f'DC: {dc}\n')
    output_text.insert(tk.END, f'Dff: {dff}\n')
    output_text.insert(tk.END, f'β: {β}\n')
    output_text.insert(tk.END, f'βD: {βd}\n')
    output_text.insert(tk.END, f'λ: {λ}\n')
    output_text.insert(tk.END, f'MTBF: {mtbf}\n')
    output_text.insert(tk.END, f'λD: {λd}\n')
    output_text.insert(tk.END, f'λDD: {λdd}\n')
    output_text.insert(tk.END, f'λDU: {λdu}\n')
    output_text.insert(tk.END, f'tDE: {tde}\n')
    output_text.insert(tk.END, f'tSE: {tse}\n')
    output_text.insert(tk.END, f'tCE: {tce}\n')
    output_text.insert(tk.END, f'PFD: {pfd}\n')

    # Output Frame
output_frame = ttk.LabelFrame(main_frame, text='Output')
output_frame.grid(row=2, column=0, padx=10, pady=10, sticky='w')

output_text = tk.Text(output_frame, height=10, width=50)
output_text.grid(row=0, column=0, padx=5, pady=5)

def add_component():

    global serial_number

    # Ask the user to input the serial number for the component
    user_serial_number = simpledialog.askstring("Serial Number", "Enter the Serial Number for the component:")
    if user_serial_number is None:
        # If the user cancels the input, do not proceed
        return

    tag_no = f"Component-{user_serial_number}"
    serial_number = user_serial_number

    tag_no = entry_tag_no.get()
    component = entry_component.get()
    make = entry_make.get()
    model = entry_model.get()
    reliability_data_ref = entry_reliability_data_ref.get()
    λsd = entry_lambda_sd.get()
    λsu = entry_lambda_su.get()
    λdd = entry_lambda_dd.get()
    λdu = entry_lambda_du.get()
    configuration = entry_configuration.get()
    proof_test_interval = entry_proof_test_interval.get()
    mttr = entry_mttr.get()
    remark= entry_remark.get()

    # Create a dictionary of the component data
    component_dict= {
        '':serial_number,
        'TAG NOS': tag_no,
        'COMPONENT/INSTRUMENT': component,
        'MAKE': make,
        'MODEL': model,
        'Reliability Data Reference': reliability_data_ref,
        'λsd': λsd,
        'λsu': λsu,
        'λdd': λdd,
        'λdu': λdu,
        'CONFIGURATION': configuration,
        'PROOF TEST INTERVAL T1 (hrs)': proof_test_interval,
        'MTTR (hrs)': mttr,
        'Remarks': remark
    }
    component_data.append(component_dict)
# Clear the input fields
    entry_tag_no.delete(0, tk.END)
    entry_component.delete(0, tk.END)
    entry_make.delete(0, tk.END)
    entry_model.delete(0, tk.END)
    entry_reliability_data_ref.delete(0, tk.END)
    entry_lambda_sd.delete(0, tk.END)
    entry_lambda_su.delete(0, tk.END)
    entry_lambda_dd.delete(0, tk.END)
    entry_lambda_du.delete(0, tk.END)
    entry_configuration.delete(0, tk.END)
    entry_proof_test_interval.delete(0, tk.END)
    entry_mttr.delete(0, tk.END)
    entry_remark.delete(0, tk.END)
    messagebox.showinfo('Success', 'Component added successfully.')
    components_added = True

    result = messagebox.askyesno("Continue", "Do you want to add more components?")
    if not result:
        # If the user chooses not to add more components, generate the report
        generate_report()
    
add_component_button = ttk.Button(main_frame, text='ADD COMPONENT', command=add_component)
add_component_button.grid(row=3, column=0, pady=5)

def generate_report():
    if len(component_data) == 0:
        messagebox.showwarning('Warning', 'No component data found. Please add components first.')
        return
   
    # Choose an existing Excel workbook
    file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx')])
    if not file_path:
        return

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Choose the sheet to write the data
    sheet_name = 'Calculations'
    if not sheet_name:
        return

    # Get the selected sheet
    sheet = workbook[sheet_name]

    # Find the starting row for writing the component data
    next_row = 2
    while sheet.cell(row=next_row, column=2).value:
        next_row += 1

    # Check for duplicate serial numbers
    existing_serial_numbers = [sheet.cell(row=row, column=1).value for row in range(2, next_row)]
    new_serial_numbers = [component[''] for component in component_data]
    duplicate_serial_numbers = set(existing_serial_numbers) & set(new_serial_numbers)

    if duplicate_serial_numbers:
        messagebox.showerror('Error', f'Duplicate serial numbers found: {", ".join(duplicate_serial_numbers)}. '
                                      f'Restart the program')
        sys.exit()
        return

    # Check if component data already exists in the sheet
    if next_row > 2:
        # Find the last row with component data
        last_row = next_row - 1

        # Check if the component data matches the existing data
        for i, component in enumerate(component_data):
            row_data = [
                sheet.cell(row=last_row, column=2).value,
                sheet.cell(row=last_row, column=3).value,
                sheet.cell(row=last_row, column=4).value,
                sheet.cell(row=last_row, column=5).value,
                sheet.cell(row=last_row, column=6).value,
                sheet.cell(row=last_row, column=7).value,
                sheet.cell(row=last_row, column=8).value,
                sheet.cell(row=last_row, column=9).value,
                sheet.cell(row=last_row, column=10).value,
                sheet.cell(row=last_row, column=11).value,
                sheet.cell(row=last_row, column=12).value,
                sheet.cell(row=last_row, column=13).value,
            ]

            if row_data == [
                component['TAG NOS'],
                component['COMPONENT/INSTRUMENT'],
                component['MAKE'],
                component['MODEL'],
                component['Reliability Data Reference'],
                component['λsd'],
                component['λsu'],
                component['λdd'],
                component['λdu'],
                component['CONFIGURATION'],
                component['PROOF TEST INTERVAL T1 (hrs)'],
                component['MTTR (hrs)'],
                component['Remarks'],
            ]:
                # Skip adding duplicate component data
                continue

            # If the component data does not match, increment the next row
            next_row += i
            break
        
    next_row = 2
    while sheet.cell(row=next_row, column=2).value:
        next_row += 1
    for component in component_data:
        sheet.cell(row=next_row, column=1).value = component['']
        sheet.cell(row=next_row, column=2).value = component['TAG NOS']
        sheet.cell(row=next_row, column=3).value = component['COMPONENT/INSTRUMENT']
        sheet.cell(row=next_row, column=4).value = component['MAKE']
        sheet.cell(row=next_row, column=5).value = component['MODEL']
        sheet.cell(row=next_row, column=6).value = component['Reliability Data Reference']
        sheet.cell(row=next_row, column=7).value = component['λsd']
        sheet.cell(row=next_row, column=8).value = component['λsu']
        sheet.cell(row=next_row, column=9).value = component['λdd']
        sheet.cell(row=next_row, column=10).value = component['λdu']
        sheet.cell(row=next_row, column=11).value = component['CONFIGURATION']
        sheet.cell(row=next_row, column=12).value = component['PROOF TEST INTERVAL T1 (hrs)']
        sheet.cell(row=next_row, column=13).value = component['MTTR (hrs)']

        # Perform calculations and write the calculated values
        try:
            λdu = float(component['λdu']) if component['λdu'] is not None else 0
            λsd = float(component['λsd']) if component['λsd'] is not None else 0
            λsu = float(component['λsu']) if component['λsu'] is not None else 0
            λdd = float(component['λdd']) if component['λdd'] is not None else 0
        except ValueError:
            messagebox.showwarning('Warning', 'Invalid input for λdu, λsd, λsu, λdd, or λdu. Please enter valid floating-point numbers.')
            return
        
        sff = (1 - λdu / (λsd + λsu + λdd + λdu)) * 100
        dc = λdd / (λdd + λdu)
        dff = 1 - sff / 100
        β = 0.1
        βd = 0.05
        λ = λsd + λsu + λdd + λdu
        mtbf = 1 / λ
        λd = λ * dff
        λdd = λd * dc
        λdu = λd * (1 - dc)
        tde = (λdu * ((float(component['PROOF TEST INTERVAL T1 (hrs)']) / 2) + float(component['MTTR (hrs)']))) + \
              (λdd * float(component['MTTR (hrs)'])) / λd
        tse = (λdu * ((float(component['PROOF TEST INTERVAL T1 (hrs)']) / 3) + float(component['MTTR (hrs)']))) + \
              (λdd * float(component['MTTR (hrs)'])) / λd
        tce = None
        pfd = (6 * ((1 - float(βd)) * float(λdd) + (1 - float(β)) * float(λdu)) ** 2 * tde * tse) + \
                            (float(βd) * float(λdd) * float(component['MTTR (hrs)'])) + \
                            (float(β) * float(λdu) * ((float(component['PROOF TEST INTERVAL T1 (hrs)']) / 2) + float(component['MTTR (hrs)'])))
        
        sheet.cell(row=next_row, column=14).value = sff
        sheet.cell(row=next_row, column=15).value = dc
        sheet.cell(row=next_row, column=16).value = dff
        sheet.cell(row=next_row, column=17).value = β
        sheet.cell(row=next_row, column=18).value = βd
        sheet.cell(row=next_row, column=19).value = λ
        sheet.cell(row=next_row, column=20).value = mtbf
        sheet.cell(row=next_row, column=21).value = λd
        sheet.cell(row=next_row, column=22).value = λdd
        sheet.cell(row=next_row, column=23).value = λdu
        sheet.cell(row=next_row, column=24).value = tde
        sheet.cell(row=next_row, column=25).value = tse
        sheet.cell(row=next_row, column=26).value = tce
        sheet.cell(row=next_row, column=27).value = pfd
        sheet.cell(row=next_row, column=29).value= component['Remarks']

        next_row += 1

    # Save the workbook
    workbook.save(file_path)
    messagebox.showinfo('Success', 'Report updated successfully.')

    window.destroy()
generate_report_button.configure(command=generate_report)
generate_report_button.grid(row=4, column=0, pady=1)

copyright_label = ttk.Label(window, text='Made By: Shantanu Ranjan(GCET).', font=('Helvetica', 8))
copyright_label.pack(side=tk.BOTTOM)



canvas.configure(scrollregion=canvas.bbox('all'))

window.mainloop()

#Second Code:
import sys
import tkinter as tk
import tkinter.filedialog as filedialog
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkinter.simpledialog import askinteger
import openpyxl

result = messagebox.askyesno("Continue", "Do you want to go to SIL Verification?")
if not result:
        # If the user chooses not to add more components, generate the report
    sys.exit()

def check_duplicates(sheet, serial_number, sif_number):
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row[0] == serial_number:
            messagebox.showerror("Error", f"Serial number '{serial_number}' already exists.")
            sys.exit()
        if row[6] == sif_number:
            messagebox.showerror("Error", f"SIF number '{sif_number}' already exists.")
            sys.exit()
# Function to handle the button click event and store the inputs in a file
def generate_report():
    # Get user inputs
    initiator_serial_number= initiator_serial_entry.get()
    initiator_tag_number = initiator_tag_entry.get()
    initiator_pid_number = initiator_pid_entry.get()
    initiator_config = initiator_config_entry.get()
    initiator_interval = initiator_interval_entry.get()
    initiator_pfd = initiator_pfd_entry.get()

    logic_sif = logic_sif_entry.get()
    logic_config = logic_config_entry.get()
    logic_interval = logic_interval_entry.get()
    logic_pfd = logic_pfd_entry.get()

    fce_tag_number = fce_tag_entry.get()
    fce_pid_number = fce_pid_entry.get()
    fce_config = fce_config_entry.get()
    fce_interval = fce_interval_entry.get()
    fce_pfd = fce_pfd_entry.get()

    total_achieved_pfd = float(initiator_pfd) + float(logic_pfd) + float(fce_pfd)
    target_pfd = target_pfd_entry.get()
    target_sil = target_sil_entry.get()
    sil_status = sil_status_entry.get()

  

    # Open file dialog to choose the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        return

    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the SIL Verification
    sheet = workbook["SIL Verification"]
 # Check for duplicate serial number and SIF number
    check_duplicates(sheet, initiator_serial_number, logic_sif)
    next_row = 4
    while sheet.cell(row=next_row, column=2).value:
        next_row += 1


    # Check if component data already exists in the sheet
    if next_row > 2:
        # Find the last row with component data
        last_row = next_row - 1


    # Write the values in the specified columns
    sheet.cell(row=next_row, column=1, value=initiator_serial_number)
    sheet.cell(row=next_row, column=2, value=initiator_tag_number)
    sheet.cell(row=next_row, column=3, value=initiator_pid_number)
    sheet.cell(row=next_row, column=4, value=initiator_config)
    sheet.cell(row=next_row, column=5, value=initiator_interval)
    sheet.cell(row=next_row, column=6, value=initiator_pfd)
    
    sheet.cell(row=next_row, column=7, value=logic_sif)
    sheet.cell(row=next_row, column=8, value=logic_config)
    sheet.cell(row=next_row, column=9, value=logic_interval)
    sheet.cell(row=next_row, column=10, value=logic_pfd)

    sheet.cell(row=next_row, column=11, value=fce_tag_number)
    sheet.cell(row=next_row, column=12, value=fce_pid_number)
    sheet.cell(row=next_row, column=13, value=fce_config)
    sheet.cell(row=next_row, column=14, value=fce_interval)
    sheet.cell(row=next_row, column=15, value=fce_pfd)

    sheet.cell(row=next_row, column=16, value=total_achieved_pfd)
    sheet.cell(row=next_row, column=17, value=target_pfd)
    sheet.cell(row=next_row, column=18, value=target_sil)
    sheet.cell(row=next_row, column=19, value=sil_status)

   
    # Save the workbook
    workbook.save(file_path)
    messagebox.showinfo('Success', 'Report updated successfully.')

    result = messagebox.askquestion("Continue?", "Do you want to add more values?", icon='info')
    if result == 'no':
        window.destroy()


    # Clear the input fields after writing to the file
    initiator_serial_entry.delete(0, tk.END)
    initiator_tag_entry.delete(0, tk.END)
    initiator_pid_entry.delete(0, tk.END)
    initiator_config_entry.delete(0, tk.END)
    initiator_interval_entry.delete(0, tk.END)
    initiator_pfd_entry.delete(0, tk.END)
    
    logic_sif_entry.delete(0, tk.END)
    logic_config_entry.delete(0, tk.END)
    logic_interval_entry.delete(0, tk.END)
    logic_pfd_entry.delete(0, tk.END)

    fce_tag_entry.delete(0, tk.END)
    fce_pid_entry.delete(0, tk.END)
    fce_config_entry.delete(0, tk.END)
    fce_interval_entry.delete(0, tk.END)
    fce_pfd_entry.delete(0, tk.END)
    target_pfd_entry.delete(0, tk.END)
    target_sil_entry.delete(0, tk.END)
    sil_status_entry.delete(0, tk.END)


# Create the GUI window

window = tk.Tk()
window.title("SIL Verification")
window.geometry("1010x400")

# Initiator Process inputs
initiator_label = tk.Label(window, text="Initiator (Process sensor)")
initiator_label.grid(row=0, column=0, sticky=tk.W)

initiator_serial_label = tk.Label(window, text="S No.")
initiator_serial_label.grid(row=1, column=0, sticky=tk.W)
initiator_serial_entry = tk.Entry(window)
initiator_serial_entry.grid(row=1, column=1)
initiator_tag_label = tk.Label(window, text="Tag number")
initiator_tag_label.grid(row=1, column=2, sticky=tk.W)
initiator_tag_entry = tk.Entry(window)
initiator_tag_entry.grid(row=1, column=3)

initiator_pid_label = tk.Label(window, text="PID no.")
initiator_pid_label.grid(row=1, column=4, sticky=tk.W)
initiator_pid_entry = tk.Entry(window)
initiator_pid_entry.grid(row=1, column=5)

initiator_config_label = tk.Label(window, text="Configuration")
initiator_config_label.grid(row=2, column=0, sticky=tk.W)
initiator_config_entry = tk.Entry(window)
initiator_config_entry.grid(row=2, column=1)

initiator_interval_label = tk.Label(window, text="Test Interval (Hrs)    ")
initiator_interval_label.grid(row=2, column=2, sticky=tk.W)
initiator_interval_entry = tk.Entry(window)
initiator_interval_entry.grid(row=2, column=3)

initiator_pfd_label = tk.Label(window, text="PFDavg    ")
initiator_pfd_label.grid(row=2, column=4, sticky=tk.W)
initiator_pfd_entry = tk.Entry(window)
initiator_pfd_entry.grid(row=2, column=5)

# Logic Solver inputs
logic_label = tk.Label(window, text="Logic Solver")
logic_label.grid(row=3, column=0, sticky=tk.W)

logic_sif_label = tk.Label(window, text="SIF No.")
logic_sif_label.grid(row=4, column=0, sticky=tk.W)
logic_sif_entry = tk.Entry(window)
logic_sif_entry.grid(row=4, column=1)
logic_config_label = tk.Label(window, text="Configuration")
logic_config_label.grid(row=4, column=2, sticky=tk.W)
logic_config_entry = tk.Entry(window)
logic_config_entry.grid(row=4, column=3)

logic_interval_label = tk.Label(window, text="Test Interval (Hrs)    ")
logic_interval_label.grid(row=4, column=4, sticky=tk.W)
logic_interval_entry = tk.Entry(window)
logic_interval_entry.grid(row=4, column=5)

logic_pfd_label = tk.Label(window, text="PFDavg    ")
logic_pfd_label.grid(row=4, column=6, sticky=tk.W)
logic_pfd_entry = tk.Entry(window)
logic_pfd_entry.grid(row=4, column=7)

# Final Control Element inputs
fce_label = tk.Label(window, text="Final Control Element")
fce_label.grid(row=5, column=0, sticky=tk.W)

fce_tag_label = tk.Label(window, text="Tag number")
fce_tag_label.grid(row=6, column=0, sticky=tk.W)
fce_tag_entry = tk.Entry(window)
fce_tag_entry.grid(row=6, column=1)

fce_pid_label = tk.Label(window, text="PID no.")
fce_pid_label.grid(row=6, column=2, sticky=tk.W)
fce_pid_entry = tk.Entry(window)
fce_pid_entry.grid(row=6, column=3)

fce_config_label = tk.Label(window, text="Configuration")
fce_config_label.grid(row=7, column=0, sticky=tk.W)
fce_config_entry = tk.Entry(window)
fce_config_entry.grid(row=7, column=1)

fce_interval_label = tk.Label(window, text="Test Interval (Hrs)")
fce_interval_label.grid(row=7, column=2, sticky=tk.W)
fce_interval_entry = tk.Entry(window)
fce_interval_entry.grid(row=7, column=3)

fce_pfd_label = tk.Label(window, text="PFDavg   ")
fce_pfd_label.grid(row=7, column=4, sticky=tk.W)
fce_pfd_entry = tk.Entry(window)
fce_pfd_entry.grid(row=7, column=5)

# Target PFD input
target_pfd_label = tk.Label(window, text="Target PFD")
target_pfd_label.grid(row=9, column=0, sticky=tk.W)
target_pfd_entry = tk.Entry(window)
target_pfd_entry.grid(row=9, column=1)

# Target SIL Level input
target_sil_label = tk.Label(window, text="Target SIL Level")
target_sil_label.grid(row=10, column=0, sticky=tk.W)
target_sil_entry = tk.Entry(window)
target_sil_entry.grid(row=10, column=1)

# SIL Status input
sil_status_label = tk.Label(window, text="SIL Status (1-Pass, 0-Fail)   ")
sil_status_label.grid(row=11, column=0, sticky=tk.W)
sil_status_entry = tk.Entry(window)
sil_status_entry.grid(row=11, column=1)

last_loaded_row = 5 
# Load Annexure Button
def load_annexure():
    global last_loaded_row
    file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx'), ('All Files', '*.*')])
    if file_path:
        selected_row = askinteger("Select Row", "Enter the row number to load data from :", minvalue=5)
        if selected_row is not None:
            annexure_data = load_annexure_data(file_path, selected_row)
            if annexure_data:
                last_loaded_row = selected_row + 1
                populate_input_fields(annexure_data)
                messagebox.showinfo('Annexure Loaded', 'Annexure data loaded successfully.')
            else:
                messagebox.showwarning('Invalid Row', 'Invalid row number. Please select a valid row from the annexure.')

load_annexure_button = ttk.Button(window, text='LOAD ANNEXURE', command=load_annexure)
load_annexure_button.grid(row=14, column=0, pady=10)

# Load Annexure Data Function
def load_annexure_data(file_path, row_number):
    annexure_data = {}
    workbook = openpyxl.load_workbook(file_path)
    sheet_name = 'Annexure-2'
    if not sheet_name:
        return

    # Get the selected sheet
    sheet = workbook[sheet_name]

    # Check if the sheet has data starting from row 4
    if sheet.max_row >= row_number:
        annexure_data['Tag number 1'] = sheet.cell(row=row_number, column=2).value
        annexure_data['PID no. 1'] = sheet.cell(row=row_number, column=3).value
        annexure_data['Configuration 1'] = sheet.cell(row=row_number, column=4).value
        annexure_data['Test Interval (Hrs) 1'] = sheet.cell(row=row_number, column=5).value
        annexure_data['PFDavg 1'] = sheet.cell(row=row_number, column=6 ).value
        annexure_data['Configuration 2'] = sheet.cell(row=row_number, column=8).value
        annexure_data['Test Interval (Hrs) 2'] = sheet.cell(row=row_number, column=9).value
        annexure_data['PFDavg 2'] = sheet.cell(row=row_number, column=10 ).value
        annexure_data['Tag number 3'] = sheet.cell(row=row_number, column=11).value
        annexure_data['PID no. 3'] = sheet.cell(row=row_number, column=12).value
        annexure_data['Configuration 3'] = sheet.cell(row=row_number, column=13).value
        annexure_data['Test Interval (Hrs) 3'] = sheet.cell(row=row_number, column=14).value
        annexure_data['PFDavg 3'] = sheet.cell(row=row_number, column=15 ).value
        annexure_data['Total Achieved PFD'] = sheet.cell(row=row_number, column=16 ).value   
    else:
        return None
    return annexure_data


# Populate Input Fields Function
def populate_input_fields(annexure_data):
    initiator_serial_entry.delete(0,tk.END)
    initiator_serial_entry.insert(0, annexure_data.get('S No.', ''))
    initiator_tag_entry.delete(0, tk.END)
    initiator_tag_entry.insert(0, annexure_data.get('Tag number 1', ''))
    initiator_pid_entry.delete(0, tk.END)
    initiator_pid_entry.insert(0, annexure_data.get('PID no. 1', ''))
    initiator_config_entry.delete(0, tk.END)
    initiator_config_entry.insert(0, annexure_data.get('Configuration 1', ''))
    initiator_interval_entry.delete(0, tk.END)
    initiator_interval_entry.insert(0, annexure_data.get('Test Interval (Hrs) 1', ''))
    initiator_pfd_entry.delete(0, tk.END)
    initiator_pfd_entry.insert(0, annexure_data.get('PFDavg 1', ''))
    logic_sif_entry.delete(0, tk.END)
    logic_sif_entry.insert(0, annexure_data.get('SIF No.', ''))
    logic_config_entry.delete(0, tk.END)
    logic_config_entry.insert(0, annexure_data.get('Configuration 2', ''))
    logic_interval_entry.delete(0, tk.END)
    logic_interval_entry.insert(0, annexure_data.get('Test Interval (Hrs) 2', ''))
    logic_pfd_entry.delete(0, tk.END)
    logic_pfd_entry.insert(0, annexure_data.get('PFDavg 2', ''))
    fce_tag_entry.delete(0, tk.END)
    fce_tag_entry.insert(0, annexure_data.get('Tag number 3', ''))
    fce_pid_entry.delete(0, tk.END)
    fce_pid_entry.insert(0, annexure_data.get('PID no. 3', ''))
    fce_config_entry.delete(0, tk.END)
    fce_config_entry.insert(0, annexure_data.get('Configuration 3', ''))
    fce_interval_entry.delete(0, tk.END)
    fce_interval_entry.insert(0, annexure_data.get('Test Interval (Hrs) 3', ''))
    fce_pfd_entry.delete(0, tk.END)
    fce_pfd_entry.insert(0, annexure_data.get('PFDavg 3', ''))

load_annexure_button = ttk.Button(window, text='LOAD ANNEXURE', command=load_annexure)
load_annexure_button.grid(row=14, column=0, pady=10)



# Generate Report button
generate_button = ttk.Button(window, text="UPDATE REPORT", command=generate_report)
generate_button.grid(row=12, column=0, pady=10)

copyright_label = ttk.Label(
    window,
    text='Made By: Shantanu Ranjan(GCET).',
    font=('Helvetica', 8)
)
# Position the copyright label at the bottom of the main frame using grid
# The 'sticky' parameter expands the label to fill the entire bottom row
copyright_label.grid(row=15, column=8, padx=10, pady=5, sticky='se')

# Start the GUI event loop
window.mainloop()

#Third Code
import sys
import tkinter as tk
import tkinter.filedialog as filedialog
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkinter.simpledialog import askinteger
import openpyxl

result = messagebox.askyesno("Continue", "Do you want to go to SIL Verification Summary?")
if not result:
        # If the user chooses not to add more components, generate the report
    sys.exit()

def check_duplicates(sheet, sif_number):
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row[0] == sif_number:
            messagebox.showerror("Error", f"SIF number '{sif_number}' already exists.")
            sys.exit()
# Function to handle the button click event and sto
# Function to handle the button click event and store the inputs in a file
def generate_report():
    # Get user inputs
    initiator_tag_number = initiator_tag_entry.get()
    initiator_pid_number = initiator_pid_entry.get()
    initiator_config = initiator_config_entry.get()
    initiator_interval = initiator_interval_entry.get()
    initiator_pfd = initiator_pfd_entry.get()

    logic_sif = logic_sif_entry.get()
    logic_config = logic_config_entry.get()
    logic_interval = logic_interval_entry.get()
    logic_pfd = logic_pfd_entry.get()

    fce_tag_number = fce_tag_entry.get()
    fce_pid_number = fce_pid_entry.get()
    fce_config = fce_config_entry.get()
    fce_interval = fce_interval_entry.get()
    fce_pfd = fce_pfd_entry.get()

    total_achieved_pfd = float(initiator_pfd) + float(logic_pfd) + float(fce_pfd)
    target_pfd = target_pfd_entry.get()
    target_sil = target_sil_entry.get()


    # Open file dialog to choose the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if not file_path:
        return

    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the SIL Verification
    sheet = workbook["Verification Summary"]
 # Check for duplicate serial number and SIF number
    check_duplicates(sheet, logic_sif)
    next_row = 4
    while sheet.cell(row=next_row, column=2).value:
        next_row += 1


    # Check if component data already exists in the sheet
    if next_row > 2:
        # Find the last row with component data
        last_row = next_row - 1


    # Write the values in the specified columns
    sheet.cell(row=next_row, column=1, value=logic_sif)
    sheet.cell(row=next_row, column=2, value=initiator_tag_number)
    sheet.cell(row=next_row, column=3, value=initiator_pid_number)
    sheet.cell(row=next_row, column=4, value=initiator_config)
    sheet.cell(row=next_row, column=5, value=initiator_interval)
    sheet.cell(row=next_row, column=6, value=initiator_pfd)

    sheet.cell(row=next_row, column=8, value=logic_config)
    sheet.cell(row=next_row, column=9, value=logic_interval)
    sheet.cell(row=next_row, column=10, value=logic_pfd)

    sheet.cell(row=next_row, column=11, value=fce_tag_number)
    sheet.cell(row=next_row, column=12, value=fce_pid_number)
    sheet.cell(row=next_row, column=13, value=fce_config)
    sheet.cell(row=next_row, column=14, value=fce_interval)
    sheet.cell(row=next_row, column=15, value=fce_pfd)

    sheet.cell(row=next_row, column=16, value=total_achieved_pfd)
    sheet.cell(row=next_row, column=17, value=target_pfd)
    sheet.cell(row=next_row, column=18, value=target_sil)

    # Save the workbook
    workbook.save(file_path)
    messagebox.showinfo('Success', 'Report updated successfully.')

    result = messagebox.askquestion("Continue?", "Do you want to add more values?", icon='info')
    if result == 'no':
        window.destroy()


    # Clear the input fields after writing to the file
    logic_sif_entry.delete(0, tk.END)
    initiator_tag_entry.delete(0, tk.END)
    initiator_pid_entry.delete(0, tk.END)
    initiator_config_entry.delete(0, tk.END)
    initiator_interval_entry.delete(0, tk.END)
    initiator_pfd_entry.delete(0, tk.END)

    logic_config_entry.delete(0, tk.END)
    logic_interval_entry.delete(0, tk.END)
    logic_pfd_entry.delete(0, tk.END)

    fce_tag_entry.delete(0, tk.END)
    fce_pid_entry.delete(0, tk.END)
    fce_config_entry.delete(0, tk.END)
    fce_interval_entry.delete(0, tk.END)
    fce_pfd_entry.delete(0, tk.END)
    target_pfd_entry.delete(0, tk.END)
    target_sil_entry.delete(0, tk.END)

# Create the GUI window

window = tk.Tk()
window.title("Verification Summary")
window.geometry("710x400")


# Initiator Process inputs
initiator_label = tk.Label(window, text="Initiator (Process sensor)")
initiator_label.grid(row=0, column=0, sticky=tk.W)

logic_sif_label = tk.Label(window, text="SIF No.")
logic_sif_label.grid(row=11, column=0, sticky=tk.W)
logic_sif_entry = tk.Entry(window)
logic_sif_entry.grid(row=11, column=1)

initiator_tag_label = tk.Label(window, text="Tag number")
initiator_tag_label.grid(row=1, column=0, sticky=tk.W)
initiator_tag_entry = tk.Entry(window)
initiator_tag_entry.grid(row=1, column=1)

initiator_pid_label = tk.Label(window, text="PID no.")
initiator_pid_label.grid(row=1, column=2, sticky=tk.W)
initiator_pid_entry = tk.Entry(window)
initiator_pid_entry.grid(row=1, column=3)

initiator_config_label = tk.Label(window, text="Configuration")
initiator_config_label.grid(row=2, column=0, sticky=tk.W)
initiator_config_entry = tk.Entry(window)
initiator_config_entry.grid(row=2, column=1)

initiator_interval_label = tk.Label(window, text="Test Interval (Hrs)    ")
initiator_interval_label.grid(row=2, column=2, sticky=tk.W)
initiator_interval_entry = tk.Entry(window)
initiator_interval_entry.grid(row=2, column=3)

initiator_pfd_label = tk.Label(window, text="PFDavg    ")
initiator_pfd_label.grid(row=2, column=4, sticky=tk.W)
initiator_pfd_entry = tk.Entry(window)
initiator_pfd_entry.grid(row=2, column=5)

# Logic Solver inputs
logic_label = tk.Label(window, text="Logic Solver")
logic_label.grid(row=3, column=0, sticky=tk.W)

logic_config_label = tk.Label(window, text="Configuration")
logic_config_label.grid(row=4, column=0, sticky=tk.W)
logic_config_entry = tk.Entry(window)
logic_config_entry.grid(row=4, column=1)

logic_interval_label = tk.Label(window, text="Test Interval (Hrs)    ")
logic_interval_label.grid(row=4, column=2, sticky=tk.W)
logic_interval_entry = tk.Entry(window)
logic_interval_entry.grid(row=4, column=3)

logic_pfd_label = tk.Label(window, text="PFDavg    ")
logic_pfd_label.grid(row=4, column=4, sticky=tk.W)
logic_pfd_entry = tk.Entry(window)
logic_pfd_entry.grid(row=4, column=5)

# Final Control Element inputs
fce_label = tk.Label(window, text="Final Control Element")
fce_label.grid(row=5, column=0, sticky=tk.W)

fce_tag_label = tk.Label(window, text="Tag number")
fce_tag_label.grid(row=6, column=0, sticky=tk.W)
fce_tag_entry = tk.Entry(window)
fce_tag_entry.grid(row=6, column=1)

fce_pid_label = tk.Label(window, text="PID no.")
fce_pid_label.grid(row=6, column=2, sticky=tk.W)
fce_pid_entry = tk.Entry(window)
fce_pid_entry.grid(row=6, column=3)

fce_config_label = tk.Label(window, text="Configuration")
fce_config_label.grid(row=7, column=0, sticky=tk.W)
fce_config_entry = tk.Entry(window)
fce_config_entry.grid(row=7, column=1)

fce_interval_label = tk.Label(window, text="Test Interval (Hrs)")
fce_interval_label.grid(row=7, column=2, sticky=tk.W)
fce_interval_entry = tk.Entry(window)
fce_interval_entry.grid(row=7, column=3)

fce_pfd_label = tk.Label(window, text="PFDavg   ")
fce_pfd_label.grid(row=7, column=4, sticky=tk.W)
fce_pfd_entry = tk.Entry(window)
fce_pfd_entry.grid(row=7, column=5)

# Target PFD input
target_pfd_label = tk.Label(window, text="Required PFD")
target_pfd_label.grid(row=9, column=0, sticky=tk.W)
target_pfd_entry = tk.Entry(window)
target_pfd_entry.grid(row=9, column=1)

# Target SIL Level input
target_sil_label = tk.Label(window, text="SIL Status")
target_sil_label.grid(row=10, column=0, sticky=tk.W)
target_sil_entry = tk.Entry(window)
target_sil_entry.grid(row=10, column=1)

last_loaded_row = 5 
# Load Annexure Button
def load_annexure():
    global last_loaded_row
    file_path = filedialog.askopenfilename(filetypes=[('Excel Files', '*.xlsx'), ('All Files', '*.*')])
    if file_path:
        selected_row = askinteger("Select Row", "Enter the row number to load data from:", minvalue=5)
        if selected_row is not None:
            annexure_data = load_annexure_data(file_path, selected_row)
            if annexure_data:
                last_loaded_row = selected_row + 1
                populate_input_fields(annexure_data)
                messagebox.showinfo('Annexure Loaded', 'Annexure data loaded successfully.')
            else:
                messagebox.showwarning('Invalid Row', 'Invalid row number. Please select a valid row from the annexure.')
load_annexure_button = ttk.Button(window, text='LOAD ANNEXURE', command=load_annexure)
load_annexure_button.grid(row=13, column=0, pady=10)

# Load Annexure Data Function
def load_annexure_data(file_path, row_number):
    annexure_data = {}
    workbook = openpyxl.load_workbook(file_path)
    sheet_name = 'Annexure-2'
    if not sheet_name:
        return

    # Get the selected sheet
    sheet = workbook[sheet_name]

    # Check if the sheet has data starting from row 4
    if sheet.max_row >= row_number:
        annexure_data['Tag number 1'] = sheet.cell(row=row_number, column=2).value
        annexure_data['PID no. 1'] = sheet.cell(row=row_number, column=3).value
        annexure_data['Configuration 1'] = sheet.cell(row=row_number, column=4).value
        annexure_data['Test Interval (Hrs) 1'] = sheet.cell(row=row_number, column=5).value
        annexure_data['PFDavg 1'] = sheet.cell(row=row_number, column=6 ).value
        annexure_data['Configuration 2'] = sheet.cell(row=row_number, column=8).value
        annexure_data['Test Interval (Hrs) 2'] = sheet.cell(row=row_number, column=9).value
        annexure_data['PFDavg 2'] = sheet.cell(row=row_number, column=10 ).value
        annexure_data['Tag number 3'] = sheet.cell(row=row_number, column=11).value
        annexure_data['PID no. 3'] = sheet.cell(row=row_number, column=12).value
        annexure_data['Configuration 3'] = sheet.cell(row=row_number, column=13).value
        annexure_data['Test Interval (Hrs) 3'] = sheet.cell(row=row_number, column=14).value
        annexure_data['PFDavg 3'] = sheet.cell(row=row_number, column=15 ).value
        annexure_data['Total Achieved PFD'] = sheet.cell(row=row_number, column=16 ).value
        return annexure_data

# Populate Input Fields Function
def populate_input_fields(annexure_data):
    initiator_tag_entry.delete(0, tk.END)
    initiator_tag_entry.insert(0, annexure_data.get('Tag number 1', ''))
    initiator_pid_entry.delete(0, tk.END)
    initiator_pid_entry.insert(0, annexure_data.get('PID no. 1', ''))
    initiator_config_entry.delete(0, tk.END)
    initiator_config_entry.insert(0, annexure_data.get('Configuration 1', ''))
    initiator_interval_entry.delete(0, tk.END)
    initiator_interval_entry.insert(0, annexure_data.get('Test Interval (Hrs) 1', ''))
    initiator_pfd_entry.delete(0, tk.END)
    initiator_pfd_entry.insert(0, annexure_data.get('PFDavg 1', ''))
    logic_config_entry.delete(0, tk.END)
    logic_config_entry.insert(0, annexure_data.get('Configuration 2', ''))
    logic_interval_entry.delete(0, tk.END)
    logic_interval_entry.insert(0, annexure_data.get('Test Interval (Hrs) 2', ''))
    logic_pfd_entry.delete(0, tk.END)
    logic_pfd_entry.insert(0, annexure_data.get('PFDavg 2', ''))
    fce_tag_entry.delete(0, tk.END)
    fce_tag_entry.insert(0, annexure_data.get('Tag number 3', ''))
    fce_pid_entry.delete(0, tk.END)
    fce_pid_entry.insert(0, annexure_data.get('PID no. 3', ''))
    fce_config_entry.delete(0, tk.END)
    fce_config_entry.insert(0, annexure_data.get('Configuration 3', ''))
    fce_interval_entry.delete(0, tk.END)
    fce_interval_entry.insert(0, annexure_data.get('Test Interval (Hrs) 3', ''))
    fce_pfd_entry.delete(0, tk.END)
    fce_pfd_entry.insert(0, annexure_data.get('PFDavg 3', ''))
    

    load_annexure_button = ttk.Button(window, text='LOAD ANNEXURE', command=load_annexure)
load_annexure_button.grid(row=14, column=0, pady=0)

# Generate Report button
generate_button = ttk.Button(window, text="UPDATE REPORT", command=generate_report)
generate_button.grid(row=12, column=0, pady=10)

copyright_label = ttk.Label(
    window,
    text='Made By: Shantanu Ranjan(GCET).',
    font=('Helvetica', 8)
)
# Position the copyright label at the bottom of the main frame using grid
# The 'sticky' parameter expands the label to fill the entire bottom row
copyright_label.grid(row=15, column=6, padx=10, pady=5, sticky='se')

# Start the GUI event loop
window.mainloop()
#END